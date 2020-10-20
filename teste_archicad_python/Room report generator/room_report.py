from archicad import ACConnection, handle_dependencies
import os, sys, uuid

handle_dependencies('openpyxl')

from openpyxl import Workbook, load_workbook

conn = ACConnection.connect()
assert conn

acc = conn.commands
act = conn.types
acu = conn.utilities

scriptFolder = os.path.dirname(os.path.realpath(__file__))

################################ CONFIGURATION #################################
outputFolder = scriptFolder
outputFileName = "Room Report.xlsx"
templateFolder = scriptFolder
templateFileName = "RDS template.xlsx"
cellAddressPropertyUserIdTable = {
    "C2": act.BuiltInPropertyUserId("Zone_ZoneName"),
    "G2": act.BuiltInPropertyUserId("Zone_ZoneNumber"),
    "G3": act.BuiltInPropertyUserId("Zone_ZoneCategoryCode"),
    "G6": act.BuiltInPropertyUserId("Zone_NetArea"),
    "G7": act.BuiltInPropertyUserId("General_NetVolume"),
    "G4": act.UserDefinedPropertyUserId(["WINDOW RATE (Expression)", "Window rate calculated"]),
    "G15": act.UserDefinedPropertyUserId(["ZONES", "Temperature Requirement"]),
    "G16": act.UserDefinedPropertyUserId(["ZONES", "Illuminance Requirement"])
}
classificationSystemName = "ARCHICAD Classification"
insertClassificationTo = "C4"
insertRelatedZonesTo = ["C" + str(row) for row in range(6, 16)]
insertEquipmentNamesTo = ["B" + str(row) for row in range(20, 57)]
insertEquipmentQuantitiesTo = ["D" + str(row) for row in range(20, 57)]
insertOpeningNamesTo = ["F" + str(row) for row in range(20, 57)]
insertOpeningElementIDsTo = ["G" + str(row) for row in range(20, 57)]
################################################################################


def getElementsClassificationDictionary(elements):
    classificationIdObjects = acc.GetClassificationsOfElements(
        elements, [acu.FindClassificationSystem(classificationSystemName)])

    def unwrapId(classification):
        if classification.classificationIds[0].classificationId.classificationItemId:
            return classification.classificationIds[0].classificationId.classificationItemId
        else:
            return act.ClassificationItemId(uuid.uuid1())

    classificationItemIds = [unwrapId(c) for c in classificationIdObjects]
    classificationDetails = acc.GetDetailsOfClassificationItems(classificationItemIds)

    def unwrapDetail(details):
        if hasattr(details, "error"):
            return "<Unclassified>"
        return details.classificationItem.id

    return dict(zip(elements, [unwrapDetail(c) for c in classificationDetails]))

def unwrapElements(elementsWrapper):
    return elementsWrapper.elements

def getAdjacentRooms(rooms):
    rawBoundaryObjects = acc.GetElementsRelatedToZones(rooms, ["Wall"])
    boundaryObjects = dict(zip(rooms, list(map(unwrapElements, rawBoundaryObjects))))

    def getGuid(elementIdArrayItem: act.ElementIdArrayItem) -> str:
        return str(elementIdArrayItem.elementId.guid)

    boundaryObjectsIds = dict({k: list(map(getGuid, v))
                                for k, v in boundaryObjects.items()})
    adjacentRooms = {}
    for room1 in rooms:
        adjacentRooms[room1] = []
        for room2 in rooms:
            if room1 is not room2 and set(boundaryObjectsIds[room2]).intersection(set(boundaryObjectsIds[room1])):
                adjacentRooms[room1].append(room2)

    return adjacentRooms


def getObjectLibPartsInRooms(rooms):
    elementInRooms = acc.GetElementsRelatedToZones(rooms, ["Object"])
    roomElements = dict(zip(rooms, list(map(unwrapElements, elementInRooms))))
    libPartNamePropertyId = acu.GetBuiltInPropertyId('General_LibraryPartName')
    propertyValuesDictionary = acu.GetPropertyValuesDictionary(sum(roomElements.values(), []), [libPartNamePropertyId])
    return dict({room: [propertyValuesDictionary[e][libPartNamePropertyId] for e in roomElements[room]] for room in rooms})


def getOpeningsInRooms(rooms):
    elementInRooms = acc.GetElementsRelatedToZones(rooms, ["Door", "Window", "Skylight", "Opening"])
    roomElements = dict(zip(rooms, list(map(unwrapElements, elementInRooms))))
    libPartNamePropertyId = acu.GetBuiltInPropertyId('General_LibraryPartName')
    elementIdPropertyId = acu.GetBuiltInPropertyId('General_ElementID')
    propertyValuesDictionary = acu.GetPropertyValuesDictionary(sum(roomElements.values(), []), [libPartNamePropertyId, elementIdPropertyId])
    return dict({room: list(zip([propertyValuesDictionary[e][libPartNamePropertyId] for e in roomElements[room]],
                                [propertyValuesDictionary[e][elementIdPropertyId] for e in roomElements[room]]))
                            for room in rooms})


class WorkBookFiller:
    def __init__(self, templatePath, rooms):
        self.templatePath = templatePath
        self.rooms = rooms
        self.cellValuesForRoom = {}
        self.cellValueRangeForRoom = {}
        self.zoneNumberPropertyId = acu.GetBuiltInPropertyId('Zone_ZoneNumber')
        self.zoneNamePropertyId = acu.GetBuiltInPropertyId('Zone_ZoneName')
        self.propertyValuesDictionary = acu.GetPropertyValuesDictionary(self.rooms, [self.zoneNumberPropertyId, self.zoneNamePropertyId])
        self.rooms = sorted(self.rooms, key=lambda r: self.propertyValuesDictionary[r][self.zoneNumberPropertyId])

    def SaveWorkbook(self, outputPath):
        workbook = self._initWorkBook()
        self._fillWorkbook(workbook)
        workbook.save(outputPath)

    def InsertPropertyValuesTo(self, cellAddressPropertyIdTable):
        propertyValuesDictionary = acu.GetPropertyValuesDictionary(self.rooms, list(cellAddressPropertyIdTable.values()))
        for cellAddress, propertyId in cellAddressPropertyIdTable.items():
            self.cellValuesForRoom[cellAddress] = {room: valuesDictionary[propertyId] for room, valuesDictionary in propertyValuesDictionary.items() if propertyId in valuesDictionary}

    def InsertClassificationTo(self, cellAddress):
        self.cellValuesForRoom[cellAddress] = getElementsClassificationDictionary(self.rooms)

    def InsertRelatedZonesTo(self, cellAddresses):
        adjacentRooms = getAdjacentRooms(self.rooms)
        adjacentRoomIds = {}
        for k, v in adjacentRooms.items():
            adjacentRoomIds[k] = [self.propertyValuesDictionary[item][self.zoneNumberPropertyId] + " - " + self.propertyValuesDictionary[item][self.zoneNamePropertyId] for item in v]

        self.cellValueRangeForRoom[tuple(cellAddresses)] = adjacentRoomIds

    def InsertObjectLibPartsTo(self, namesCellAddresses, countsCellAddresses):
        libpartsInRooms = getObjectLibPartsInRooms(self.rooms)
        libpartNamesInRooms = {}
        libpartCountsInRooms = {}
        for k, v in libpartsInRooms.items():
            libpartNamesInRooms[k] = sorted(list(set(v)))
            libpartCountsInRooms[k] = [v.count(libpartName) for libpartName in libpartNamesInRooms[k]]

        self.cellValueRangeForRoom[tuple(namesCellAddresses)] = libpartNamesInRooms
        self.cellValueRangeForRoom[tuple(countsCellAddresses)] = libpartCountsInRooms

    def InsertOpeningsTo(self, namesCellAddresses, idsCellAddresses):
        openingsInRooms = getOpeningsInRooms(self.rooms)
        libpartNamesInRooms = {}
        elementIdsInRooms = {}
        for k, v in openingsInRooms.items():
            openings = sorted(v, key=lambda t: t[1])
            libpartNamesInRooms[k] = [t[0] for t in openings]
            elementIdsInRooms[k] = [t[1] for t in openings]

        self.cellValueRangeForRoom[tuple(namesCellAddresses)] = libpartNamesInRooms
        self.cellValueRangeForRoom[tuple(idsCellAddresses)] = elementIdsInRooms

    def _initWorkBook(self):
        workbook = load_workbook(self.templatePath)
        workbook.template = False
        return workbook

    def _fillWorkbook(self, workbook):
        base = workbook.active
        for room in self.rooms:
            roomName = self.propertyValuesDictionary[room][self.zoneNamePropertyId].replace("/", "-")
            roomId = self.propertyValuesDictionary[room][self.zoneNumberPropertyId].replace("/", "-")
            if self.propertyValuesDictionary[room][self.zoneNumberPropertyId]:
                worksheet = workbook.copy_worksheet(base)
                worksheet.title = f"{roomId} {roomName}"
                for cellAddress, cellValueForRoom in self.cellValuesForRoom.items():
                    if room in cellValueForRoom:
                        worksheet[cellAddress] = cellValueForRoom[room]
                        print(f"{worksheet.title}!{cellAddress}={cellValueForRoom[room]}")

                for cellAddressrange, cellvalue in self.cellValueRangeForRoom.items():
                    for cellAddress, value in zip(cellAddressrange, cellvalue[room]):
                        worksheet[cellAddress] = value
                        print(f"{worksheet.title}!{cellAddress}={value}")

        workbook.remove(base)


templatePath = os.path.join(templateFolder, templateFileName)
wbFiller = WorkBookFiller(templatePath, acc.GetElementsByType("Zone"))

wbFiller.InsertPropertyValuesTo(dict(zip(
    list(cellAddressPropertyUserIdTable.keys()),
    acc.GetPropertyIds(list(cellAddressPropertyUserIdTable.values()))
)))

wbFiller.InsertRelatedZonesTo(insertRelatedZonesTo)
wbFiller.InsertObjectLibPartsTo(insertEquipmentNamesTo, insertEquipmentQuantitiesTo)
wbFiller.InsertOpeningsTo(insertOpeningNamesTo, insertOpeningElementIDsTo)
wbFiller.InsertClassificationTo(insertClassificationTo)

outputPath = os.path.join(outputFolder, outputFileName)
wbFiller.SaveWorkbook(outputPath)
acu.OpenFile(outputPath)

if os.path.exists(outputPath):
    print("Saved Room Report")
