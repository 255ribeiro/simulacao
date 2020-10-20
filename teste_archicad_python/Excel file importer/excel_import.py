from archicad import ACConnection, handle_dependencies
from typing import List, Dict, Any
import os, sys, uuid

handle_dependencies('openpyxl')

from openpyxl import load_workbook

conn = ACConnection.connect()
assert conn

acc = conn.commands
act = conn.types
acu = conn.utilities

scriptFolder = os.path.dirname(os.path.realpath(__file__))

################################ CONFIGURATION #################################
inputFolder = scriptFolder
inputFileName = "BeamAndWallGeometry.xlsx"
################################################################################


excelFilePath = os.path.join(inputFolder, inputFileName)
wb = load_workbook(excelFilePath)
elemPropertyValues = []
for sheet in wb.worksheets:
	maxCol = sheet.max_column
	maxRow = sheet.max_row
	newPropertyValues = {}
	elementIds = []
	propertyIds = []
	for col in range (2, maxCol + 1):
		propertyIds.append(act.PropertyId(uuid.UUID(sheet.cell(1, col).value)))
	for row in range (3, maxRow + 1):
		newPropertyValues[row-3] = {col-2: sheet.cell(row, col).value for col in range (2, maxCol + 1)}
		elementIds.append(act.ElementId(uuid.UUID(sheet.cell(row, 1).value)))

	propertyValuesOfElements = acc.GetPropertyValuesOfElements(elementIds, propertyIds)

	for ii in range(len(newPropertyValues)):
		for jj in range(len(newPropertyValues[ii])):
			try:
				propertyValue = propertyValuesOfElements[ii].propertyValues[jj].propertyValue
				propertyValue.value = newPropertyValues[ii][jj]
				propertyValue.status = "normal"
				elemPropertyValues.append(act.ElementPropertyValue(elementIds[ii], propertyIds[jj], propertyValue))
			except:
				continue

acc.SetPropertyValuesOfElements(elemPropertyValues)	

# Print the result
elementIds = [i.elementId for i in elemPropertyValues]
propertyIds = [act.PropertyId(guid) for guid in set(i.propertyId.guid for i in elemPropertyValues)]
propertyValuesDictionary = acu.GetPropertyValuesDictionary(elementIds, propertyIds)
for elementId, valuesDictionary in propertyValuesDictionary.items():
	for propertyId, value in valuesDictionary.items():
		print(f"{elementId.guid} {propertyId.guid} {value}")
