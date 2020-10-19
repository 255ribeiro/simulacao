from archicad import ACConnection, handle_dependencies
from typing import List
import os, sys

handle_dependencies('openpyxl')

from openpyxl import Workbook

conn = ACConnection.connect()
assert conn

acc = conn.commands
act = conn.types
acu = conn.utilities

scriptFolder = os.path.dirname(os.path.realpath(__file__))

################################ CONFIGURATION #################################
worksheetTitlesAndElements = {
    "Beams": acc.GetElementsByType("Beam"),
    "Walls": acc.GetElementsByType("Wall")
}
propertyUserIds = [
    act.BuiltInPropertyUserId("General_ElementID"),
    act.BuiltInPropertyUserId("General_Height"),
    act.BuiltInPropertyUserId("General_Width"),
    act.BuiltInPropertyUserId("General_Thickness")
]
outputFolder = scriptFolder
outputFileName = "BeamAndWallGeometry.xlsx"
################################################################################


def AutoFitWorksheetColumns(ws):
    for columnCells in ws.columns:
        length = max(len(str(cell.value)) for cell in columnCells)
        ws.column_dimensions[columnCells[0].column_letter].width = length


def PrintWorksheetContent(ws):
    for columnCells in ws.columns:
        for cell in columnCells:
            print(f"{ws.title}!{cell.column_letter}{cell.row}={cell.value}")


def FillExcelWorksheetWithPropertyValuesOfElements(ws, propertyIds: List[act.PropertyIdArrayItem], elements: List[act.ElementIdArrayItem]):
    propertyValuesDictionary = acu.GetPropertyValuesDictionary(elements, propertyIds)
    propertyDefinitionsDictionary = dict(zip(propertyIds, acc.GetDetailsOfProperties(propertyIds)))

    ws.cell(row=2, column=1).value = "Element Guid"
    row = 3
    for element, valuesDictionary in propertyValuesDictionary.items():
        ws.cell(row=row, column=1).value = str(element.elementId.guid)
        column = 2
        for propertyId, propertyValue in valuesDictionary.items():
            if row == 3:
                ws.cell(row=1, column=column).value = str(propertyId.propertyId.guid)
                propertyDefinition = propertyDefinitionsDictionary[propertyId].propertyDefinition
                ws.cell(row=2, column=column).value = f"{propertyDefinition.group.name} / {propertyDefinition.name}"
            ws.cell(row=row, column=column).value = propertyValue
            column += 1
        row += 1

    AutoFitWorksheetColumns(ws)
    PrintWorksheetContent(ws)


propertyIds = acc.GetPropertyIds(propertyUserIds)
wb = Workbook()
ws = wb.active

i = 0
for title, elements in worksheetTitlesAndElements.items():
    if i == 0:
        ws.title = title
    else:
        ws = wb.create_sheet(title)
    FillExcelWorksheetWithPropertyValuesOfElements(ws, propertyIds, elements)
    i += 1

excelFilePath = os.path.join(outputFolder, outputFileName)
wb.save(excelFilePath)
acu.OpenFile(excelFilePath)

if os.path.exists(excelFilePath):
    print("Saved Excel")
